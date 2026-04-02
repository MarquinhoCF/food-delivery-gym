"""
generate_table.py

Gera automaticamente a planilha de resultados (objective_table.xlsx) a partir
dos dados produzidos pelo script evaluate_agents.py.

- Descobre agentes e modelos PPO varrendo os diretórios de resultados
- Tenta carregar metrics_data.npz; se não encontrar, tenta metrics_data.json
- Usa SimulationStats para acessar os dados agregados
- Constrói o Excel dinamicamente: sem mapeamentos manuais de colunas
- Replica o estilo visual do template original
- Destaca em negrito o melhor agente por cenário/objetivo em cada aba
"""

import json
import os
import argparse
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from food_delivery_gym.main.environment.food_delivery_gym_env import FoodDeliveryGymEnv
from food_delivery_gym.main.scenarios import get_all_scenarios, get_defaults_scenarios

# ── Configuração de diretórios ────────────────────────────────────────────────

DEFAULT_RESULTS_DIR = "./data/runs/execucoes"
DEFAULT_OUTPUT_PATH = "./data/teste/runs/tabelas/objective_table.xlsx"
ALL_OBJECTIVES      = FoodDeliveryGymEnv.REWARD_OBJECTIVES
ALL_SCENARIOS       = get_all_scenarios()
DEFAULT_SCENARIOS   = get_defaults_scenarios()
SCENARIO_LABELS     = {"initial": "Inicial", "simple": "Simples", "medium": "Médio", "medium_driver_cap_4": "Médio (Cap. 4)", "complex": "Complexo"}
METRICS             = ["avg", "std_dev", "median", "mode"]
METRIC_LABELS       = ["Média", "Desvio Padrão", "Mediana", "Moda"]
ROWS_PER_OBJECTIVE  = len(METRICS)   # 4 linhas por objetivo
HEADER_ROWS         = 2              # linhas de cabeçalho antes dos dados

# Heurísticas conhecidas: dir_name → label legível
KNOWN_HEURISTICS = {
    "random":                    "Motorista Aleatório",
    "first_driver":              "Primeiro Motorista",
    "nearest_driver":            "Motorista mais Próximo",
    "lowest_route_cost":         "Motorista de Menor Custo de Rota",
    "lowest_marginal_route_cost":"Motorista de Menor Custo Marginal de Rota",
}

# Chaves de SimulationStats.aggregate → nome da aba
# Deve corresponder ao que finalize() grava em self.aggregate
AGG_KEY_TO_SHEET = {
    "rewards":       "Recompensas",
    "delivery_time": "Tempo Efetivo Gasto",
    "distance":      "Distância Percorrida",
}

# Recompensas: maior é melhor; demais métricas: menor é melhor
HIGHER_IS_BETTER = {"Recompensas"}

# ── Paleta de cores (extraída do template original) ───────────────────────────
COLOR_HEADER_BG = "104862"   # azul escuro — cabeçalhos
COLOR_HEADER_FG = "FFFFFF"   # branco
COLOR_OBJ_BG    = "CAEEFB"   # azul claro — célula do objetivo
COLOR_ROW_ALT_A = "A6CAEC"   # azul médio — média / mediana
COLOR_ROW_ALT_B = "DCEAF7"   # azul pálido — desvio padrão / moda
COLOR_SEP_BG    = "104862"   # separador entre cenários

ROW_FILLS = [COLOR_ROW_ALT_A, COLOR_ROW_ALT_B, COLOR_ROW_ALT_A, COLOR_ROW_ALT_B]

# ── Descrições dos objetivos (preservadas do template) ───────────────────────
OBJECTIVE_DESCRIPTIONS = {
    1:  "Minimizar o tempo de entrega a partir da expectativa de tempo gasto com a entrega a cada passo",
    2:  "Minimizar o custo de operação a partir da expectativa da distância a ser percorrida a cada passo",
    3:  "Minimizar o tempo de entrega dos motoristas a partir do tempo efetivo gasto a cada passo",
    4:  "Minimizar o custo de operação a partir da distância efetiva a cada passo",
    5:  "Minimizar o tempo de entrega a partir da expectativa de tempo gasto com a entrega ao final do episódio",
    6:  "Minimizar o custo de operação a partir da expectativa da distância a ser percorrida ao final do episódio",
    7:  "Minimizar o tempo de entrega dos motoristas a partir do tempo efetivo gasto ao final do episódio",
    8:  "Minimizar o custo de operação a partir da distância efetiva ao final do episódio",
    9:  "Minimizar o tempo de entrega dos motoristas a partir do tempo efetivo gasto (Com penalização 5x para pedidos não coletados)  a cada passo",
    10: "Minimizar o tempo de entrega dos motoristas a partir do tempo efetivo gasto (Com penalização 5x para pedidos não coletados) ao final do episódio",
    11: "Maximizar o número de pedidos entregues Recompensa positiva a cada pedido entregue",
    12: "Penaliza pelo tempo total de cada pedido entregue neste step. Quanto mais rápido o pedido for entregue, menor a penalidade (maior a recompensa)",
    13: "Penaliza pelo tempo gasto de cada pedido que está na pipeline de entrega (pedidos prontos e não entregues) e de cada pedido entregue neste step. ",
}

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border(style="medium") -> Border:
    s = Side(border_style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_border() -> Border:
    s = Side(border_style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(cell, label):
    cell.value     = label
    cell.font      = _font(bold=True, color=COLOR_HEADER_FG)
    cell.fill      = _fill(COLOR_HEADER_BG)
    cell.border    = _border("medium")
    cell.alignment = _center()

def style_separator(cell, label=""):
    cell.value     = label
    cell.font      = _font(bold=True, color=COLOR_HEADER_FG)
    cell.fill      = _fill(COLOR_SEP_BG)
    cell.border    = _border("medium")
    cell.alignment = _center()

def style_obj_label(cell, label):
    cell.value     = label
    cell.font      = _font(bold=True)
    cell.fill      = _fill(COLOR_OBJ_BG)
    cell.border    = _thin_border()
    cell.alignment = _center()

def style_obj_desc(cell, label):
    cell.value     = label
    cell.font      = _font()
    cell.fill      = _fill(COLOR_OBJ_BG)
    cell.border    = _thin_border()
    cell.alignment = _center()

def style_metric_label(cell, label, row_idx):
    cell.value     = label
    cell.font      = _font(bold=True)
    cell.fill      = _fill(ROW_FILLS[row_idx])
    cell.border    = _thin_border()
    cell.alignment = _center()

def style_data_cell(cell, value, row_idx, bold=False):
    cell.value     = value
    cell.font      = _font(bold=bold)
    cell.fill      = _fill(ROW_FILLS[row_idx])
    cell.border    = _thin_border()
    cell.alignment = _center()

# ── Carregamento direto de agregados (NPZ ou JSON) ───────────────────────────

def _load_aggregate_npz(npz_path: str) -> dict:
    """
    Lê apenas as chaves agg__<metric>__<stat> do arquivo NPZ.

    Retorna dict no formato: { "rewards": { "avg": 1.2, ... }, ... }
    """
    aggregate: dict = {}
    with np.load(npz_path, allow_pickle=False) as raw:
        for key in raw.files:
            parts = key.split("__")
            if parts[0] == "agg" and len(parts) == 3:
                _, metric, stat = parts
                val = raw[key]
                aggregate.setdefault(metric, {})[stat] = (
                    int(val[0]) if stat == "n" else float(val[0])
                )
    return aggregate


def _load_aggregate_json(json_path: str) -> dict:
    """
    Lê o campo 'aggregate' do arquivo JSON.

    Retorna dict no formato: { "rewards": { "avg": 1.2, ... }, ... }
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("aggregate", {})


def load_aggregate(agent_dir: str) -> dict | None:
    """
    Carrega os agregados do diretório do agente.

    Ordem de tentativa:
      1. metrics_data.npz  →  chaves agg__*
      2. metrics_data.json →  campo "aggregate"

    Retorna None se nenhum arquivo for encontrado ou ocorrer erro.
    """
    npz_path  = os.path.join(agent_dir, "metrics_data.npz")
    json_path = os.path.join(agent_dir, "metrics_data.json")

    if os.path.exists(npz_path):
        try:
            return _load_aggregate_npz(npz_path)
        except Exception as e:
            print(f"  [Erro NPZ] {npz_path}: {e}")

    if os.path.exists(json_path):
        try:
            return _load_aggregate_json(json_path)
        except Exception as e:
            print(f"  [Erro JSON] {json_path}: {e}")

    return None


def _has_metrics_file(agent_dir: str) -> bool:
    """Verifica se há ao menos um arquivo de métricas no diretório do agente."""
    return (
        os.path.isfile(os.path.join(agent_dir, "metrics_data.npz")) or
        os.path.isfile(os.path.join(agent_dir, "metrics_data.json"))
    )

# ── Descoberta de agentes ─────────────────────────────────────────────────────

def agent_label(dir_name: str) -> str:
    """Converte nome de diretório em label legível."""
    if dir_name in KNOWN_HEURISTICS:
        return KNOWN_HEURISTICS[dir_name]
    if dir_name.startswith("ppo_"):
        suffix = dir_name[4:]   # remove "ppo_"
        return f"PPO — {suffix}"
    return dir_name


def discover_agents(results_dir: str, objectives: list, scenarios: list) -> list:
    """
    Varre results_dir para descobrir todos os agentes presentes.

    Um agente é válido se seu diretório contém metrics_data.npz ou
    metrics_data.json. Retorna lista ordenada: heurísticas conhecidas
    primeiro (na ordem de KNOWN_HEURISTICS), depois modelos PPO
    em ordem alfabética.
    """
    found = set()
    for obj in objectives:
        for scenario in scenarios:
            path = os.path.join(results_dir, f"obj_{obj}", f"{scenario}_scenario")
            if not os.path.isdir(path):
                continue
            for entry in os.scandir(path):
                if entry.is_dir() and _has_metrics_file(entry.path):
                    found.add(entry.name)

    heuristics = [k for k in KNOWN_HEURISTICS if k in found]
    ppo_models = sorted(d for d in found if d not in KNOWN_HEURISTICS)
    return heuristics + ppo_models

# ── Construção do Excel ───────────────────────────────────────────────────────

def build_workbook(results_dir: str, objectives: list, scenarios: list, agents: list) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for agg_key, sheet_name in AGG_KEY_TO_SHEET.items():
        ws = wb.create_sheet(sheet_name)
        _build_sheet(ws, sheet_name, agg_key, results_dir, objectives, scenarios, agents)

    return wb


def _build_sheet(ws, sheet_name: str, agg_key: str, results_dir: str,
                 objectives: list, scenarios: list, agents: list):
    """Constrói uma aba completa."""

    # ── Dimensões dinâmicas ──────────────────────────────────────────────────
    # Layout de colunas por cenário:
    #   col 1: "Objetivo N"
    #   col 2: descrição
    #   Para cada cenário i:
    #     sep_col          = 3 + i * (1 + len(agents))
    #     first_agent_col  = sep_col + 1
    #     last_agent_col   = sep_col + len(agents)

    n = len(agents)

    def scenario_start(i):
        """Coluna do separador de estatísticas para o cenário i (0-indexed)."""
        return 3 + i * (n + 1)

    def first_agent_col(i):
        return scenario_start(i) + 1

    # ── Linha 1: Cenário ─────────────────────────────────────────────────────
    style_header(ws.cell(1, 1), "Cenário")
    style_separator(ws.cell(1, 2), "//////////")

    for i, scenario in enumerate(scenarios):
        sep         = scenario_start(i)
        label_start = sep
        label_end   = sep + n
        style_header(ws.cell(1, label_start), SCENARIO_LABELS[scenario])
        if label_end > label_start:
            ws.merge_cells(
                start_row=1, start_column=label_start,
                end_row=1,   end_column=label_end
            )

    # ── Linha 2: Cabeçalhos dos agentes ──────────────────────────────────────
    style_header(ws.cell(2, 1), "Heurísticas")
    style_header(ws.cell(2, 2), "Descrição")

    for sc_idx, _ in enumerate(scenarios):
        sep = scenario_start(sc_idx)
        style_header(ws.cell(2, sep), "Estatísticas")
        for j, agent in enumerate(agents):
            style_header(ws.cell(2, first_agent_col(sc_idx) + j), agent_label(agent))

    # ── Linhas de dados (objetivos × 4 métricas) ─────────────────────────────
    for obj_i, obj in enumerate(objectives):
        base_row = HEADER_ROWS + 1 + obj_i * ROWS_PER_OBJECTIVE

        for m_i, (metric_key, metric_label) in enumerate(zip(METRICS, METRIC_LABELS)):
            row = base_row + m_i

            # Coluna A: label do objetivo (merge nas 4 linhas do objetivo)
            if m_i == 0:
                style_obj_label(ws.cell(row, 1), f"Objetivo {obj}")
                if ROWS_PER_OBJECTIVE > 1:
                    ws.merge_cells(
                        start_row=row, start_column=1,
                        end_row=row + ROWS_PER_OBJECTIVE - 1, end_column=1
                    )
                # Coluna B: descrição (merge nas 4 linhas do objetivo)
                desc = OBJECTIVE_DESCRIPTIONS.get(obj, "")
                style_obj_desc(ws.cell(row, 2), desc)
                ws.merge_cells(
                    start_row=row, start_column=2,
                    end_row=row + ROWS_PER_OBJECTIVE - 1, end_column=2
                )
            else:
                c_a = ws.cell(row, 1)
                c_a.fill   = _fill(COLOR_OBJ_BG)
                c_a.border = _thin_border()
                c_b = ws.cell(row, 2)
                c_b.fill   = _fill(COLOR_OBJ_BG)
                c_b.border = _thin_border()

            # Por cenário
            for sc_i, scenario in enumerate(scenarios):
                sep = scenario_start(sc_i)

                # Rótulo estatístico
                style_metric_label(ws.cell(row, sep), metric_label, m_i)

                # Dados de cada agente via SimulationStats
                for j, agent in enumerate(agents):
                    agent_dir = os.path.join(
                        results_dir, f"obj_{obj}", f"{scenario}_scenario", agent
                    )

                    value = _get_metric_value(agent_dir, agg_key, metric_key)
                    style_data_cell(ws.cell(row, first_agent_col(sc_i) + j), value, m_i)

    # ── Destacar melhor média por cenário/objetivo ────────────────────────────
    _highlight_best(ws, sheet_name, objectives, scenarios, agents,
                    scenario_start, first_agent_col, n)

    # ── Altura das linhas ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 45

    # ── Larguras de coluna ────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 48
    for i, _ in enumerate(scenarios):
        sep = scenario_start(i)
        ws.column_dimensions[get_column_letter(sep)].width = 14
        for j in range(n):
            col   = first_agent_col(i) + j
            label = agent_label(agents[j])
            ws.column_dimensions[get_column_letter(col)].width = max(18, min(len(label) * 1.1, 40))

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "B3"


def _get_metric_value(agent_dir: str, agg_key: str, metric_key: str) -> float | None:
    """
    Carrega os agregados do diretório do agente e retorna o valor de
    aggregate[agg_key][metric_key], ou None se indisponível.

    agg_key   : chave do agregado ("rewards", "delivery_time", "distance")
    metric_key: estatística desejada ("avg", "std_dev", "median", "mode")
    """
    aggregate = load_aggregate(agent_dir)
    if aggregate is None:
        return None

    value = aggregate.get(agg_key, {}).get(metric_key)

    if isinstance(value, (np.floating, np.integer)):
        return float(value)
    if isinstance(value, float) and value != value:   # NaN
        return None
    return value


def _highlight_best(ws, sheet_name, objectives, scenarios, agents,
                    scenario_start_fn, first_agent_col_fn, n):
    """Aplica negrito nas células da melhor média por cenário/objetivo."""
    higher_is_better = sheet_name in HIGHER_IS_BETTER

    for obj_i, obj in enumerate(objectives):
        avg_row = HEADER_ROWS + 1 + obj_i * ROWS_PER_OBJECTIVE  # linha da média

        for sc_i in range(len(scenarios)):
            best_val = None
            best_col = None

            for j in range(n):
                col   = first_agent_col_fn(sc_i) + j
                value = ws.cell(avg_row, col).value
                try:
                    val = float(value)
                except (TypeError, ValueError):
                    continue

                if best_val is None:
                    best_val = val
                    best_col = col
                elif higher_is_better and val > best_val:
                    best_val = val
                    best_col = col
                elif not higher_is_better and val < best_val:
                    best_val = val
                    best_col = col

            if best_col is not None:
                for m_i in range(ROWS_PER_OBJECTIVE):
                    cell = ws.cell(avg_row + m_i, best_col)
                    cell.font = _font(bold=True)

# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Gera a planilha de resultados a partir dos dados do evaluate_agents.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--results-dir", "-r",
        default=DEFAULT_RESULTS_DIR,
        help=f"Diretório raiz com os resultados (obj_N/). Padrão: {DEFAULT_RESULTS_DIR}",
    )
    parser.add_argument(
        "--output", "-out",
        default=DEFAULT_OUTPUT_PATH,
        help=f"Caminho do arquivo Excel de saída. Padrão: {DEFAULT_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--objectives", "-o",
        nargs="+", type=int, default=ALL_OBJECTIVES, metavar="N",
        help="Objetivos a incluir. Padrão: todos (1–13).",
    )
    parser.add_argument(
        "--scenarios", "-s",
        nargs="+", choices=ALL_SCENARIOS, default=DEFAULT_SCENARIOS, metavar="SCENARIO",
        help=f"Cenários a incluir. Opções: {DEFAULT_SCENARIOS}. Padrão: todos.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    print(f"Varrendo diretório: {args.results_dir}")
    agents = discover_agents(args.results_dir, args.objectives, args.scenarios)

    if not agents:
        print(
            "[AVISO] Nenhum agente encontrado. Verifique o --results-dir e se os "
            "arquivos metrics_data.npz ou metrics_data.json existem."
        )
        return

    print(f"Agentes encontrados ({len(agents)}): {agents}")
    print(f"Objetivos: {args.objectives}")
    print(f"Cenários:  {args.scenarios}")

    wb = build_workbook(args.results_dir, args.objectives, args.scenarios, agents)
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"\nPlanilha salva em: {args.output}")


if __name__ == "__main__":
    main()